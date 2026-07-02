"""Génère data_dictionary.xlsx (BI) et access_control_matrix.xlsx (Cyber)."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))

HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def styliser(ws, n_cols, n_rows):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for r in range(2, n_rows + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER


# --- 1) Dictionnaire de données --------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Dictionnaire"
ws.append(["Colonne", "Type", "Description", "Exemple", "Valeurs manquantes ?"])
dico = [
    ["id_commande", "entier", "Identifiant unique de la commande", "100123", "Non"],
    ["date", "date", "Date de la commande (AAAA-MM-JJ)", "2025-11-15", "Non"],
    ["id_client", "texte", "Identifiant du client", "CLI0042", "Non"],
    ["produit", "texte", "Nom du produit commandé", "Casque Bluetooth Pro", "Non"],
    ["categorie", "texte", "Catégorie du produit", "Electronique", "Non"],
    ["quantite", "entier", "Nombre d'unités commandées (> 0)", "2", "Non"],
    ["prix_unitaire", "décimal", "Prix catalogue unitaire (€)", "129", "Non"],
    ["remise", "décimal", "Taux de remise appliqué (0 à 0,2)", "0.1", "Non"],
    ["montant", "décimal", "Montant total ligne = prix x qté x (1-remise)", "232.20", "Non"],
    ["region", "texte", "Région de livraison", "Ile-de-France", "Remplacé par 'Non renseigne'"],
    ["canal", "texte", "Canal de vente", "Web / Mobile / Magasin", "Non"],
    ["moyen_paiement", "texte", "Moyen de paiement", "Carte / PayPal / ...", "Remplacé par 'Inconnu'"],
    ["mois", "texte", "Mois de la commande (AAAA-MM), dérivé de date", "2025-11", "Non"],
]
for row in dico:
    ws.append(row)
styliser(ws, 5, len(dico) + 1)
for col, w in zip("ABCDE", [16, 10, 42, 22, 26]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# feuille méta
ws2 = wb.create_sheet("Infos_dataset")
ws2.append(["Propriété", "Valeur"])
meta = [
    ["Fichier", "data/processed/dataset_clean.csv"],
    ["Nombre de lignes", "3155"],
    ["Période couverte", "2025-01-01 au 2025-12-31"],
    ["Nettoyage appliqué", "doublons retirés, quantités <=0 retirées, montants recalculés, NA remplacés"],
    ["Anomalies résiduelles", "12 commandes (prix incohérent) laissées volontairement pour la détection"],
]
for row in meta:
    ws2.append(row)
styliser(ws2, 2, len(meta) + 1)
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 60

wb.save(os.path.join(BASE, "data_dictionary.xlsx"))
print("data_dictionary.xlsx créé")


# --- 2) Matrice de contrôle d'accès ----------------------------------------
wb2 = Workbook()
ws = wb2.active
ws.title = "Matrice d'accès"
ressources = ["KPI publics", "KPI détaillés", "Documents (FAQ)",
              "Données clients (PII)", "Logs / audit", "Configuration"]
ws.append(["Rôle"] + ressources)
matrice = {
    "admin":       ["✓", "✓", "✓", "✓", "✓", "✓"],
    "analyste":    ["✓", "✓", "✓", "✗", "✗", "✗"],
    "utilisateur": ["✓", "✗", "✓", "✗", "✗", "✗"],
}
for role, droits in matrice.items():
    ws.append([role] + droits)
styliser(ws, len(ressources) + 1, len(matrice) + 1)
ws.column_dimensions["A"].width = 14
for col in "BCDEFG":
    ws.column_dimensions[col].width = 16
# colorer ✓ vert / ✗ rouge
for r in range(2, len(matrice) + 2):
    for c in range(2, len(ressources) + 2):
        cell = ws.cell(row=r, column=c)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.value == "✓":
            cell.font = Font(name="Arial", bold=True, color="006100")
            cell.fill = PatternFill("solid", start_color="C6EFCE")
        else:
            cell.font = Font(name="Arial", bold=True, color="9C0006")
            cell.fill = PatternFill("solid", start_color="FFC7CE")

ws3 = wb2.create_sheet("Légende")
ws3.append(["Rôle", "Description"])
leg = [
    ["admin", "Administrateur : accès complet, y compris données clients et logs"],
    ["analyste", "Analyste métier : KPI et dashboard, pas d'accès aux données personnelles"],
    ["utilisateur", "Utilisateur simple : KPI publics et FAQ uniquement"],
]
for row in leg:
    ws3.append(row)
styliser(ws3, 2, len(leg) + 1)
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 70

wb2.save(os.path.join(BASE, "access_control_matrix.xlsx"))
print("access_control_matrix.xlsx créé")
